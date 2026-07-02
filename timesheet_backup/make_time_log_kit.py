#!/usr/bin/env python3
"""Build the "Time Log" kit workbook — the Asana-timesheet backup tabs.

Generates ``ODL Time Log (import into planning sheet).xlsx`` with four tabs
(Time Log / Read Me / Time Summary / Lists) meant to be imported into the live
"ODL Project and Capacity Planning" Google Sheet. Dropdowns use the exact
people / project / phase names the team already sees in Asana (pulled from the
estimator's nightly ``data_all`` snapshot), and the log is pre-seeded with this
year's real Asana time entries so the format is shown by example.

Re-run any time to refresh the dropdown vocabularies and the seed:
    python3 make_time_log_kit.py
"""
import csv
import datetime
import os

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))                 # "PM intern" (this file lives in odl_pm_dashboard/timesheet_backup/)
ASANA_DIR = os.environ.get("ODL_ASANA_DIR") or os.path.join(ROOT, "odl_estimator", "data_all")
OUT = os.path.join(HERE, "ODL Time Log (import into planning sheet).xlsx")

SEED_SINCE = "2026-01-01"   # seed the log with this year's real Asana entries
MAXROWS = 3000              # rows prepared with validation + helper formulas

NAVY, GOLD, GREY = "0C2340", "C99700", "F2F2F2"

# team lookup by first name — mirrors the planning workbook's roster
# (odl_pm_dashboard/build.py ROSTER_SEED); Asana uses full names.
TEAM_BY_FIRST = {
    "Yi": "Design", "Kuangchen": "Design", "Brianna": "Design", "Bri": "Design",
    "Matthew": "Media", "Kevin": "Media", "Colin": "Media", "Derrick": "Media",
    "KC": "Media", "Adam": "Media", "Tim": "Media", "Naomi": "Media",
    "Michael": "PM", "Annie": "PM", "Jordan": "PM", "Lawrence": "PM",
    "Janyl": "PM", "Sonia": "PM",
    "Nina": "Intern", "Maddie": "Intern", "Minyoung": "Intern",
}

# canonical phase choices — the section names the team already uses in Asana
PHASES = [
    "Analysis", "Design", "Design Development", "Pre-Production", "Production",
    "Post-Production", "Course Build", "QA / Launch", "Post-Project Evaluation",
    "PM Time Tracking", "Media Timetracking", "Meetings", "Other",
]


def read_csv(name):
    path = os.path.join(ASANA_DIR, name)
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def build_roster(entries, tasks):
    """People dropdown: everyone who has logged time, plus regular assignees."""
    seen, roster = set(), []

    def add(name):
        n = (name or "").strip()
        if n and "@" not in n and n.lower() not in seen:
            seen.add(n.lower())
            roster.append(n)

    for r in entries:                       # historical loggers first
        add(r.get("entry_author"))
    counts = {}
    for r in tasks:
        a = (r.get("assignee") or "").strip()
        if a:
            counts[a] = counts.get(a, 0) + 1
    for a, c in sorted(counts.items(), key=lambda kv: -kv[1]):
        if c >= 10:                         # regulars only, no one-off guests
            add(a)
    roster.sort()
    return [(n, TEAM_BY_FIRST.get(n.split()[0], "")) for n in roster]


def build_projects(projects):
    """Project dropdown: active (non-archived) Asana projects, exact names."""
    names = sorted({(r.get("project_name") or "").strip()
                    for r in projects
                    if (r.get("project_name") or "").strip()
                    and (r.get("archived") or "").strip().lower() not in ("true", "1", "yes")},
                   key=str.lower)
    return names


def seed_rows(entries):
    rows = []
    for r in entries:
        d = (r.get("entry_date") or "").strip()[:10]
        if d < SEED_SINCE or len(d) != 10:
            continue
        try:
            hrs = round(float(r.get("hours") or 0), 2)
        except ValueError:
            continue
        if not hrs:
            continue
        rows.append([d, (r.get("entry_author") or "").strip(),
                     (r.get("project_name") or "").strip(),
                     (r.get("canonical_phase") or "").strip(),
                     (r.get("task_name") or "").strip(), hrs, "Asana"])
    rows.sort(key=lambda x: x[0])
    return rows


def style_header(ws, ncols, title_row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=title_row, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=NAVY)
        cell.alignment = Alignment(vertical="center")


def main():
    entries = read_csv("time_entries.csv")
    tasks = read_csv("tasks_raw.csv")
    projects = read_csv("projects.csv")

    roster = build_roster(entries, tasks)
    proj_names = build_projects(projects)
    seeds = seed_rows(entries)
    today = datetime.date.today().isoformat()

    wb = Workbook()

    # ------------------------------------------------------------- Time Log --
    ws = wb.active
    ws.title = "Time Log"
    headers = ["Date", "Person", "Project", "Phase", "Task / note", "Hours", "Source", "Month"]
    ws.append(headers)
    style_header(ws, len(headers))
    for r in seeds:
        ws.append(r + [None])
    for ri in range(2, MAXROWS + 1):
        ws.cell(row=ri, column=8).value = f'=IF($A{ri}="","",TEXT($A{ri},"YYYY-MM"))'
        ws.cell(row=ri, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=ri, column=6).number_format = "0.0#"
    for ri in range(2, len(seeds) + 2):     # seeded dates arrive as text → real dates
        c = ws.cell(row=ri, column=1)
        c.value = datetime.date.fromisoformat(c.value) if isinstance(c.value, str) else c.value

    widths = [11, 20, 42, 22, 44, 8, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{MAXROWS}"

    n_people, n_proj, n_phase = len(roster), len(proj_names), len(PHASES)
    dvs = [
        (DataValidation(type="list", formula1=f"'Lists'!$A$2:$A${n_people + 1}",
                        allow_blank=True, showErrorMessage=False), f"B2:B{MAXROWS}"),
        (DataValidation(type="list", formula1=f"'Lists'!$D$2:$D${n_proj + 1}",
                        allow_blank=True, showErrorMessage=False), f"C2:C{MAXROWS}"),
        (DataValidation(type="list", formula1=f"'Lists'!$F$2:$F${n_phase + 1}",
                        allow_blank=True, showErrorMessage=False), f"D2:D{MAXROWS}"),
        (DataValidation(type="list", formula1="'Lists'!$H$2:$H$3",
                        allow_blank=True, showErrorMessage=False), f"G2:G{MAXROWS}"),
    ]
    for dv, rng in dvs:
        ws.add_data_validation(dv)
        dv.add(rng)

    # Asana-synced rows render grey so hand-entered work stands out
    ws.conditional_formatting.add(
        f"A2:H{MAXROWS}",
        FormulaRule(formula=['$G2="Asana"'],
                    fill=PatternFill("solid", start_color=GREY),
                    font=Font(color="808080")))

    # -------------------------------------------------------------- Read Me --
    rm = wb.create_sheet("Read Me (Time Log)")
    rm.column_dimensions["A"].width = 110
    lines = [
        ("ODL Time Log — how this works", True),
        ("", False),
        ("Asana's time-entry (timesheet) feature lapsed on June 24, 2026. Until it's back, this tab is", False),
        ("where we keep our shared picture of where project time goes — same names, same phases,", False),
        ("same feel as Asana, just in the planning sheet we already use.", False),
        ("", False),
        ("Logging time (under a minute):", True),
        ("  1. Open the Time Log tab and go to the first empty row.", False),
        ("  2. Date · your name · project · phase — each is a dropdown with the same options as Asana.", False),
        ("  3. Add a short task note if you like, then hours (e.g. 1.5). Leave Source as Manual.", False),
        ("", False),
        ("Log as you go, or batch it — a Friday 5-minute wrap-up works fine. One row per chunk of work.", False),
        ("", False),
        ("What's automatic:", True),
        ("  • Grey rows marked “Asana” are synced from Asana's records — please don't edit those.", False),
        ("  • The PM dashboard picks up Manual rows in its nightly rebuild, so logged time keeps", False),
        ("    flowing into the same charts and estimates as before. Nothing else changes.", False),
        ("  • The Time Summary tab totals hours by person, month, and project as you type.", False),
        ("", False),
        ("When Asana time tracking comes back:", True),
        ("  Just switch back to logging in Asana. Rows here stay part of the record, and the two", False),
        ("  sources are merged without double-counting.", False),
        ("", False),
        ("Questions or a project missing from the dropdown? Ping the PM team — you can also just", False),
        ("type a name that isn't in the list; the dropdowns are suggestions, not gates.", False),
        ("", False),
        (f"Kit generated {today} from the latest Asana snapshot.", False),
    ]
    for i, (txt, bold) in enumerate(lines, 1):
        c = rm.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=12 if i == 1 else 10,
                      color=NAVY if bold else "222222")
        c.alignment = Alignment(wrap_text=False)

    # --------------------------------------------------------- Time Summary --
    sm = wb.create_sheet("Time Summary")
    months = [f"2026-{m:02d}" for m in range(1, 13)]
    sm["A1"] = "Hours by person — 2026 (from the Time Log tab)"
    sm["A1"].font = Font(bold=True, size=12, color=NAVY)
    hdr = ["Person"] + months + ["Total", "Most recent entry"]
    for ci, h in enumerate(hdr, 1):
        sm.cell(row=2, column=ci, value=h)
    style_header(sm, len(hdr), title_row=2)
    r = 3
    for name, _team in roster:
        sm.cell(row=r, column=1, value=name)
        for ci, m in enumerate(months, 2):
            sm.cell(row=r, column=ci).value = (
                f"=SUMIFS('Time Log'!$F:$F,'Time Log'!$B:$B,$A{r},'Time Log'!$H:$H,{get_column_letter(ci)}$2)")
            sm.cell(row=r, column=ci).number_format = "0.0;-0.0;—"
        tc = get_column_letter(len(months) + 1)
        sm.cell(row=r, column=len(months) + 2).value = f"=SUM(B{r}:{get_column_letter(len(months)+1)}{r})"
        sm.cell(row=r, column=len(months) + 2).number_format = "0.0"
        sm.cell(row=r, column=len(months) + 3).value = (
            f'=IF(COUNTIF(\'Time Log\'!$B:$B,$A{r})=0,"—",'
            f'TEXT(MAXIFS(\'Time Log\'!$A:$A,\'Time Log\'!$B:$B,$A{r}),"yyyy-mm-dd"))')
        r += 1
    sm.column_dimensions["A"].width = 20
    for ci in range(2, len(months) + 3):
        sm.column_dimensions[get_column_letter(ci)].width = 9
    sm.column_dimensions[get_column_letter(len(months) + 3)].width = 16

    pr = r + 2
    sm.cell(row=pr, column=1, value="Hours by project (everything in the Time Log tab)").font = \
        Font(bold=True, size=12, color=NAVY)
    sm.cell(row=pr + 1, column=1, value="Project")
    sm.cell(row=pr + 1, column=2, value="Hours")
    style_header(sm, 2, title_row=pr + 1)
    sm.cell(row=pr + 2, column=1).value = f"=UNIQUE('Time Log'!$C$2:$C${MAXROWS})"
    for i in range(80):
        rr = pr + 2 + i
        sm.cell(row=rr, column=2).value = (
            f'=IF($A{rr}="","",SUMIFS(\'Time Log\'!$F:$F,\'Time Log\'!$C:$C,$A{rr}))')
        sm.cell(row=rr, column=2).number_format = "0.0"

    # ---------------------------------------------------------------- Lists --
    ls = wb.create_sheet("Lists")
    ls["A1"], ls["B1"] = "People (as in Asana)", "Team"
    ls["D1"], ls["F1"], ls["H1"] = "Projects (as in Asana)", "Phases", "Source"
    for cell in ("A1", "B1", "D1", "F1", "H1"):
        ls[cell].font = Font(bold=True, color="FFFFFF")
        ls[cell].fill = PatternFill("solid", start_color=NAVY)
    for i, (n, t) in enumerate(roster, 2):
        ls.cell(row=i, column=1, value=n)
        ls.cell(row=i, column=2, value=t)
    for i, p in enumerate(proj_names, 2):
        ls.cell(row=i, column=4, value=p)
    for i, p in enumerate(PHASES, 2):
        ls.cell(row=i, column=6, value=p)
    ls["H2"], ls["H3"] = "Manual", "Asana"
    for col, w in (("A", 22), ("B", 8), ("D", 46), ("F", 22), ("H", 10)):
        ls.column_dimensions[col].width = w

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  seeded {len(seeds)} Asana entries since {SEED_SINCE}")
    print(f"  dropdowns: {len(roster)} people, {len(proj_names)} projects, {len(PHASES)} phases")


if __name__ == "__main__":
    main()
