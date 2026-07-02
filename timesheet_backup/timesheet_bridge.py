#!/usr/bin/env python3
"""Keep the workbook Time Log tab and Asana's time records in step.

Two subcommands (both safe to re-run; nothing is ever deleted):

  sync-from-asana   Append NEW Asana time entries (from the estimator's
                    nightly ``data_all/time_entries.csv``) into the Time Log
                    tab, marked Source="Asana". Incremental: only entries
                    dated after the newest Asana row already in the sheet
                    (override with --since YYYY-MM-DD). This is the
                    "when Asana works, the workbook updates automatically"
                    direction — run it after the nightly pull, or weekly.

  export-merged     Write ``time_entries_merged.csv`` — every Asana entry
                    from the CSV plus every Manual row from the Time Log tab,
                    de-duplicated, in the same schema as time_entries.csv.
                    Drop-in feed for the estimator's calibration during the
                    Asana outage.

Target workbook: --xlsx PATH (default: the kit file next to this script).
Once the Time Log tab lives in the live Google Sheet, run sync-from-asana
against a fresh export of it (odl_pm_dashboard/fetch_drive.py downloads one),
then paste/import the updated tab back — or wire up Sheets-API write access
(service account needs editor + spreadsheets scope; see README).

The PM dashboard does NOT need this script: build.py reads Manual rows from
the workbook's Time Log tab directly on every nightly build.
"""
import argparse
import csv
import datetime
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))                 # "PM intern" (this file lives in odl_pm_dashboard/timesheet_backup/)
ASANA_DIR = os.environ.get("ODL_ASANA_DIR") or os.path.join(ROOT, "odl_estimator", "data_all")
DEFAULT_XLSX = os.path.join(HERE, "ODL Time Log (import into planning sheet).xlsx")
CSV_PATH = os.path.join(ASANA_DIR, "time_entries.csv")
MERGED_OUT = os.path.join(ASANA_DIR, "time_entries_merged.csv")

CSV_FIELDS = ["project_gid", "project_name", "task_gid", "task_name", "section",
              "canonical_phase", "assignee", "responsible_team", "entry_author",
              "entry_date", "minutes", "hours"]


def norm_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()[:10]
    return s if len(s) == 10 else ""


def key(date, person, project, task, hours):
    return (date, (person or "").strip().lower(), (project or "").strip().lower(),
            (task or "").strip().lower(), round(float(hours or 0), 2))


def load_csv_entries():
    if not os.path.exists(CSV_PATH):
        sys.exit(f"missing {CSV_PATH} — run the estimator pull first")
    rows = []
    for r in csv.DictReader(open(CSV_PATH, encoding="utf-8")):
        d = norm_date(r.get("entry_date"))
        try:
            h = round(float(r.get("hours") or 0), 2)
        except ValueError:
            continue
        if d and h:
            rows.append(r | {"entry_date": d, "hours": h})
    return rows


def open_log(path):
    from openpyxl import load_workbook
    wb = load_workbook(path)
    if "Time Log" not in wb.sheetnames:
        sys.exit(f"'{path}' has no 'Time Log' tab")
    return wb, wb["Time Log"]


def read_log(ws):
    """-> (rows, last_data_row). rows = [{row, date, person, project, phase,
    task, hours, source}] for rows with a date + hours."""
    out, last = [], 1
    for ri in range(2, ws.max_row + 1):
        vals = [ws.cell(row=ri, column=c).value for c in range(1, 8)]
        d = norm_date(vals[0])
        if any(v not in (None, "") for v in vals[:7]):
            last = ri
        try:
            h = round(float(vals[5] or 0), 2)
        except (TypeError, ValueError):
            h = 0
        if d and h:
            out.append(dict(row=ri, date=d, person=vals[1] or "", project=vals[2] or "",
                            phase=vals[3] or "", task=vals[4] or "", hours=h,
                            source=(vals[6] or "").strip() or "Manual"))
    return out, last


def cmd_sync(args):
    from collections import Counter
    csv_rows = load_csv_entries()
    wb, ws = open_log(args.xlsx)
    log, last = read_log(ws)
    # multiset, so two genuinely identical entries (same task, day, hours,
    # logged twice) both make it into the mirror
    have = Counter(key(r["date"], r["person"], r["project"], r["task"], r["hours"])
                   for r in log)
    asana_dates = [r["date"] for r in log if r["source"].lower() == "asana"]
    since = args.since or (max(asana_dates) if asana_dates else "0000-00-00")

    added = 0
    for r in sorted(csv_rows, key=lambda x: x["entry_date"]):
        if r["entry_date"] < since:
            continue
        k = key(r["entry_date"], r.get("entry_author"), r.get("project_name"),
                r.get("task_name"), r["hours"])
        if have[k] > 0:
            have[k] -= 1
            continue
        last += 1
        vals = [datetime.date.fromisoformat(r["entry_date"]), r.get("entry_author") or "",
                r.get("project_name") or "", r.get("canonical_phase") or "",
                r.get("task_name") or "", r["hours"], "Asana"]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=last, column=ci, value=v)
        ws.cell(row=last, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=last, column=8).value = f'=IF($A{last}="","",TEXT($A{last},"YYYY-MM"))'
        added += 1

    if added and not args.dry_run:
        wb.save(args.xlsx)
    print(f"sync-from-asana: {added} new Asana entries (since {since})"
          f"{' [dry-run, not saved]' if args.dry_run else ''}")


def cmd_export(args):
    csv_rows = load_csv_entries()
    wb, ws = open_log(args.xlsx)
    log, _ = read_log(ws)
    have = {key(r["entry_date"], r.get("entry_author"), r.get("project_name"),
                r.get("task_name"), r["hours"]) for r in csv_rows}
    # loose key too, so a manual row later re-logged in Asana with the task
    # named differently still collapses to one entry
    loose = {(r["entry_date"], (r.get("entry_author") or "").strip().lower(),
              (r.get("project_name") or "").strip().lower(), r["hours"]) for r in csv_rows}

    merged = [{f: r.get(f, "") for f in CSV_FIELDS} for r in csv_rows]
    added = 0
    for m in log:
        if m["source"].strip().lower() == "asana":
            continue                       # mirror rows: canonical copy is the CSV
        if key(m["date"], m["person"], m["project"], m["task"], m["hours"]) in have:
            continue
        if (m["date"], m["person"].strip().lower(), m["project"].strip().lower(), m["hours"]) in loose:
            continue
        merged.append({"project_gid": "", "project_name": m["project"], "task_gid": "",
                       "task_name": m["task"] or "Manual time log", "section": m["phase"],
                       "canonical_phase": m["phase"], "assignee": m["person"],
                       "responsible_team": "", "entry_author": m["person"],
                       "entry_date": m["date"], "minutes": round(m["hours"] * 60),
                       "hours": m["hours"]})
        added += 1
    merged.sort(key=lambda r: str(r["entry_date"]))
    out = args.out or MERGED_OUT
    if not args.dry_run:
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
            w.writeheader()
            w.writerows(merged)
    print(f"export-merged: {len(csv_rows)} Asana + {added} manual -> {out}"
          f"{' [dry-run, not written]' if args.dry_run else ''}")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("command", choices=["sync-from-asana", "export-merged"])
    ap.add_argument("--xlsx", default=DEFAULT_XLSX, help="workbook holding the Time Log tab")
    ap.add_argument("--since", help="sync entries dated on/after this (YYYY-MM-DD)")
    ap.add_argument("--out", help="output path for export-merged")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    {"sync-from-asana": cmd_sync, "export-merged": cmd_export}[args.command](args)


if __name__ == "__main__":
    main()
